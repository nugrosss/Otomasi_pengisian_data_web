from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted

# Define the folder path where images are located
image_folder_path = 'D:/magang/Coba_Selenium/Paktera/Shapes'
save_file_path = 'D:/magang/Coba_Selenium/Paktera/multiple_images.xlsx'  # Path to save the Excel file

# Create a new workbook and activate the first worksheet
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Images and Titles"

# Resize cells for images and titles
for row in range(1, 11):  # Adjust row range as needed
    worksheet.row_dimensions[row].height = 230  # Set row height for image space
    for col in range(1, 3):  # Columns A and B
        col_letter = get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = 40  # Set column width for image space

# Get the list of images with sorted order
images = natsorted(glob.glob(f'{image_folder_path}/*.png'))

# Check if there are images to insert
if not images:
    print("No images found in the specified folder.")
else:
    # Insert images and titles
    for index, image_path in enumerate(images):
        # Insert image in the first column
        img = Image(image_path)
        cell_position = 'A' + str(index + 1)
        worksheet.add_image(img, cell_position)
        
        # Insert image title in the second column
        title = os.path.basename(image_path)
        worksheet.cell(row=index + 1, column=2, value=title)
        
        # Print confirmation of each inserted image and title
        print(f"Inserted {title} at {cell_position}")

# Save the workbook to the specified path
workbook.save(save_file_path)
print(f"Workbook saved successfully at {save_file_path}")
